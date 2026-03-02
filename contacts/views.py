from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Q, Count
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Contact, ContactStats, AutomationLog
from .forms import ContactForm
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import json
from django.conf import settings


# ─────────────────────────────────────────────
#  LOGIN / LOGOUT
# ─────────────────────────────────────────────

def admin_login_view(request):
    """Login page for admin/dashboard access"""
    if request.user.is_authenticated:
        return redirect('dashboard')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            error = 'Invalid username or password.'

    return render(request, 'contacts/login.html', {'error': error})


def admin_logout_view(request):
    """Logout and redirect to landing page"""
    logout(request)
    return redirect('landing')


# ─────────────────────────────────────────────
#  PUBLIC LANDING PAGE  (clients see this at /)
# ─────────────────────────────────────────────

def landing_view(request):
    """Public-facing sign-up page where contacts submit their own details"""
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            contact = form.save()
            update_daily_stats()
            request.session['contact_number'] = contact.id
            request.session['contact_name'] = contact.full_name
            return redirect('thank-you')
    else:
        form = ContactForm()

    context = {
        'form': form,
        'total_contacts': Contact.objects.count(),
    }
    return render(request, 'contacts/landing.html', context)


def thank_you_view(request):
    """Shown after a successful sign-up"""
    context = {
        'contact_number': request.session.get('contact_number', ''),
        'contact_name': request.session.get('contact_name', ''),
        'total_contacts': Contact.objects.count(),
    }
    return render(request, 'contacts/thank_you.html', context)


# ─────────────────────────────────────────────
#  CONTACT CRUD
# ─────────────────────────────────────────────

class ContactListView(ListView):
    model = Contact
    template_name = 'contacts/contact_list.html'
    context_object_name = 'contacts'
    paginate_by = 25

    def get_queryset(self):
        queryset = Contact.objects.all()

        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(surname__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(whatsapp_contact__icontains=search_query) |
                Q(social_media_handle__icontains=search_query)
            )

        category = self.request.GET.get('category', '')
        if category:
            queryset = queryset.filter(category=category)

        country = self.request.GET.get('country', '')
        if country:
            queryset = queryset.filter(country__icontains=country)

        social_media = self.request.GET.get('social_media', '')
        if social_media:
            queryset = queryset.filter(social_media_platform=social_media)

        age_category = self.request.GET.get('age_category', '')
        if age_category:
            queryset = queryset.filter(age_category=age_category)

        sort_by = self.request.GET.get('sort', '-date_added')
        valid_sorts = ['-date_added', 'date_added', '-followers_count', 'followers_count', 'first_name', '-first_name']
        if sort_by in valid_sorts:
            queryset = queryset.order_by(sort_by)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_contacts'] = Contact.objects.count()
        context['today_contacts'] = Contact.objects.filter(
            date_added__date=timezone.now().date()
        ).count()
        context['categories'] = Contact.CATEGORY_CHOICES
        context['age_categories'] = Contact.AGE_CATEGORY_CHOICES
        context['social_media_platforms'] = Contact.SOCIAL_MEDIA_CHOICES
        context['countries'] = Contact.objects.values_list('country', flat=True).distinct().order_by('country')
        context['current_filters'] = {
            'search': self.request.GET.get('search', ''),
            'category': self.request.GET.get('category', ''),
            'country': self.request.GET.get('country', ''),
            'social_media': self.request.GET.get('social_media', ''),
            'age_category': self.request.GET.get('age_category', ''),
            'sort': self.request.GET.get('sort', '-date_added'),
        }
        return context


class ContactCreateView(CreateView):
    model = Contact
    form_class = ContactForm
    template_name = 'contacts/contact_form.html'
    success_url = reverse_lazy('contact-list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Contact {self.object.full_name} added successfully!')
        update_daily_stats()
        return response


class ContactUpdateView(UpdateView):
    model = Contact
    form_class = ContactForm
    template_name = 'contacts/contact_form.html'
    success_url = reverse_lazy('contact-list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Contact {self.object.full_name} updated successfully!')
        return response


class ContactDeleteView(DeleteView):
    model = Contact
    template_name = 'contacts/contact_confirm_delete.html'
    success_url = reverse_lazy('contact-list')

    def delete(self, request, *args, **kwargs):
        contact = self.get_object()
        messages.success(request, f'Contact {contact.full_name} deleted successfully!')
        return super().delete(request, *args, **kwargs)


# ─────────────────────────────────────────────
#  DASHBOARD  (protected - login required)
# ─────────────────────────────────────────────

@login_required(login_url='login')
def dashboard_view(request):
    """Dashboard with statistics and analytics"""

    total_contacts = Contact.objects.count()
    today_contacts = Contact.objects.filter(
        date_added__date=timezone.now().date()
    ).count()
    this_week_contacts = Contact.objects.filter(
        date_added__gte=timezone.now() - timedelta(days=7)
    ).count()

    first_contact = Contact.objects.order_by('date_added').first()
    if first_contact:
        days_since_start = (timezone.now().date() - first_contact.date_added.date()).days
    else:
        days_since_start = 0

    CATEGORY_LABELS = dict(Contact.CATEGORY_CHOICES)
    raw_category_stats = (
        Contact.objects
        .values('category')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    category_stats = [
        {
            'category': CATEGORY_LABELS.get(row['category'], row['category']),
            'count': row['count'],
        }
        for row in raw_category_stats
    ]

    country_stats = (
        Contact.objects
        .values('country')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    SOCIAL_LABELS = dict(Contact.SOCIAL_MEDIA_CHOICES)
    raw_social_stats = (
        Contact.objects
        .values('social_media_platform')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    social_stats = [
        {
            'social_media': SOCIAL_LABELS.get(row['social_media_platform'], row['social_media_platform']),
            'count': row['count'],
        }
        for row in raw_social_stats
    ]

    context = {
        'total_contacts': total_contacts,
        'today_contacts': today_contacts,
        'this_week_contacts': this_week_contacts,
        'days_since_start': days_since_start,
        'category_stats': category_stats,
        'country_stats': country_stats,
        'social_stats': social_stats,
    }
    return render(request, 'contacts/dashboard.html', context)


# ─────────────────────────────────────────────
#  EXPORTS
# ─────────────────────────────────────────────

def export_contacts_excel(request):
    contacts = Contact.objects.all()
    if request.GET.get('category'):
        contacts = contacts.filter(category=request.GET.get('category'))
    if request.GET.get('country'):
        contacts = contacts.filter(country__icontains=request.GET.get('country'))
    if request.GET.get('social_media'):
        contacts = contacts.filter(social_media_platform=request.GET.get('social_media'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SK Contacts"

    header_fill = PatternFill(start_color="38B6FF", end_color="38B6FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        'First Name', 'Surname', 'Full Name', 'WhatsApp Contact', 'Email',
        'Category', 'Age Category', 'Country', 'Country Code',
        'Social Media', 'Handle', 'Followers', 'School', 'Level',
        'Date Added', 'Days Since Added'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row, contact in enumerate(contacts, start=2):
        ws.cell(row=row, column=1, value=contact.first_name)
        ws.cell(row=row, column=2, value=contact.surname)
        ws.cell(row=row, column=3, value=contact.full_name)
        ws.cell(row=row, column=4, value=contact.whatsapp_contact)
        ws.cell(row=row, column=5, value=contact.email or '')
        ws.cell(row=row, column=6, value=contact.get_category_display())
        ws.cell(row=row, column=7, value=contact.age_category)
        ws.cell(row=row, column=8, value=contact.country)
        ws.cell(row=row, column=9, value=contact.country_code)
        ws.cell(row=row, column=10, value=contact.get_social_media_platform_display())
        ws.cell(row=row, column=11, value=contact.social_media_handle or '')
        ws.cell(row=row, column=12, value=contact.followers_count)
        ws.cell(row=row, column=13, value=contact.school or '')
        ws.cell(row=row, column=14, value=contact.level or '')
        ws.cell(row=row, column=15, value=contact.date_added.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=16, value=contact.days_since_added)

    for column in ws.columns:
        max_length = 0
        col_cells = [cell for cell in column]
        for cell in col_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_cells[0].column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'SK_Contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_contacts_csv(request):
    contacts = Contact.objects.all()
    if request.GET.get('category'):
        contacts = contacts.filter(category=request.GET.get('category'))
    if request.GET.get('country'):
        contacts = contacts.filter(country__icontains=request.GET.get('country'))
    if request.GET.get('social_media'):
        contacts = contacts.filter(social_media_platform=request.GET.get('social_media'))

    response = HttpResponse(content_type='text/csv')
    filename = f'SK_Contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        'First Name', 'Surname', 'Full Name', 'WhatsApp Contact', 'Email',
        'Category', 'Age Category', 'Country', 'Country Code',
        'Social Media', 'Handle', 'Followers', 'School', 'Level',
        'Date Added', 'Days Since Added'
    ])

    for contact in contacts:
        writer.writerow([
            contact.first_name, contact.surname, contact.full_name,
            contact.whatsapp_contact, contact.email or '',
            contact.get_category_display(), contact.age_category,
            contact.country, contact.country_code,
            contact.get_social_media_platform_display(),
            contact.social_media_handle or '', contact.followers_count,
            contact.school or '', contact.level or '',
            contact.date_added.strftime('%Y-%m-%d %H:%M'),
            contact.days_since_added,
        ])

    return response


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

def update_daily_stats():
    today = timezone.now().date()
    total_contacts = Contact.objects.count()
    today_contacts = Contact.objects.filter(date_added__date=today).count()
    ContactStats.objects.update_or_create(
        date=today,
        defaults={
            'total_contacts': total_contacts,
            'contacts_added_today': today_contacts,
        }
    )


def contact_detail_api(request, pk):
    contact = get_object_or_404(Contact, pk=pk)
    data = {
        'id': contact.id,
        'first_name': contact.first_name,
        'surname': contact.surname,
        'full_name': contact.full_name,
        'email': contact.email,
        'whatsapp_contact': contact.whatsapp_contact,
        'whatsapp_link': contact.get_whatsapp_link(),
        'category': contact.get_category_display(),
        'age_category': contact.age_category,
        'country': contact.country,
        'social_media': contact.get_social_media_platform_display(),
        'followers_count': contact.followers_count,
        'days_since_added': contact.days_since_added,
    }
    return JsonResponse(data)