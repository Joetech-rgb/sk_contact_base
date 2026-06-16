from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.db.models import Count
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from .models import APIKey, Contact, Category
from .models import Contact, Category, ReferralSource # Add ReferralSource here

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# CATEGORY ADMIN
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin.register(ReferralSource)
class ReferralSourceAdmin(admin.ModelAdmin):
    list_display = ['label', 'slug', 'is_active']
@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    """
    Admins can add, edit and remove categories here.
    These categories populate the registration form dropdown.
    """
    list_display  = ['name', 'contact_count']
    search_fields = ['name']
    ordering      = ['name']

    def contact_count(self, obj):
        count = obj.contact_set.count()
        return format_html('<strong>{}</strong>', count)
    contact_count.short_description = 'Contacts'


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# CONTACT ADMIN
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin.register(Contact)
class ContactAdmin(admin.ModelAdmin):

    # â”€â”€ List view â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    list_display = [
        'formatted_name',
        'whatsapp_with_code',
        'email',
        'category',
        'age_range',
        'country',
        'region',
        'platform_badge',
        'handle_link',
        'follower_range',
        'school_info',
        'date_added',
        'days_since_added',
    ]

    list_filter = [
        'category',
        'age_range',
        'country',
        'platform',
        'follower_range',
        'school_category',
        'date_added',
    ]

    search_fields = [
        'first_name',
        'surname',
        'email',
        'whatsapp_number',
        'handle',
        'country',
        'region',
        'school_name',
    ]

    ordering = ['-date_added']

    date_hierarchy = 'date_added'

    # â”€â”€ Detail view layout â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    fieldsets = (
        ('Basic Details', {
            'fields': ('first_name', 'surname', 'email')
        }),
        ('WhatsApp & Location', {
            'fields': ('whatsapp_number', 'country_code', 'country', 'region'),
            'description': 'WhatsApp number is stored with country code prefix.',
        }),
        ('Profile', {
            'fields': ('age_range', 'category'),
        }),
        ('Social Media', {
            'fields': ('platform', 'handle', 'follower_range'),
        }),
        ('Education (Optional)', {
            'fields': ('school_category', 'school_name', 'level_year'),
            'classes': ('collapse',),
        }),
        ('Meta', {
            'fields': ('date_added',),
            'classes': ('collapse',),
        }),
    )

    readonly_fields = ['date_added']

    # â”€â”€ Bulk actions â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    actions = [
        'export_to_excel',
        'export_to_csv',
    ]

    # â”€â”€ Custom display columns â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def formatted_name(self, obj):
        return format_html('<strong>{} {}</strong>', obj.first_name, obj.surname)
    formatted_name.short_description = 'Name'
    formatted_name.admin_order_field = 'first_name'

    def whatsapp_with_code(self, obj):
        """Show full WhatsApp number (code + number) with click-to-chat link."""
        # Build the full number for display: country_code + whatsapp_number
        code = obj.country_code or ''
        num  = obj.whatsapp_number or ''

        # If the number already starts with +, use it directly
        if num.startswith('+'):
            full = num
        else:
            full = code + num

        # Strip + and spaces for the wa.me link
        wa_num = full.replace('+', '').replace(' ', '')

        return format_html(
            '<a href="https://wa.me/{}" target="_blank" '
            'style="color:#25D366;text-decoration:none;white-space:nowrap;">'
            'ðŸ’¬ {}</a>',
            wa_num, full
        )
    whatsapp_with_code.short_description = 'WhatsApp'
    whatsapp_with_code.admin_order_field = 'whatsapp_number'

    def platform_badge(self, obj):
        icons = {
            'instagram': 'ðŸ“¸',
            'tiktok':    'ðŸŽµ',
            'facebook':  'ðŸ‘¥',
            'twitter':   'ðŸ¦',
            'youtube':   'â–¶ï¸',
            'snapchat':  'ðŸ‘»',
            'linkedin':  'ðŸ’¼',
            'other':     'ðŸŒ',
        }
        icon = icons.get(obj.platform, 'ðŸŒ')
        return format_html('{} {}', icon, obj.platform.title() if obj.platform else 'â€”')
    platform_badge.short_description = 'Platform'
    platform_badge.admin_order_field = 'platform'

    def handle_link(self, obj):
        """
        Clickable handle â€” links directly to the profile on the platform.
        Admin can toggle to the person's actual profile.
        """
        if not obj.handle:
            return 'â€”'

        handle = obj.handle.lstrip('@')

        urls = {
            'instagram': f'https://instagram.com/{handle}',
            'tiktok':    f'https://tiktok.com/@{handle}',
            'facebook':  f'https://facebook.com/{handle}',
            'twitter':   f'https://twitter.com/{handle}',
            'youtube':   f'https://youtube.com/@{handle}',
            'snapchat':  f'https://snapchat.com/add/{handle}',
            'linkedin':  f'https://linkedin.com/in/{handle}',
        }

        url = urls.get(obj.platform)
        if url:
            return format_html(
                '<a href="{}" target="_blank" style="color:#1E6FD9;">@{}</a>',
                url, handle
            )
        return format_html('@{}', handle)
    handle_link.short_description = 'Handle'

    def school_info(self, obj):
        if not obj.school_name:
            return 'â€”'
        parts = [obj.school_name]
        if obj.school_category:
            parts.append(f'({obj.get_school_category_display()})')
        if obj.level_year:
            parts.append(f'Yr {obj.level_year}')
        return ' '.join(parts)
    school_info.short_description = 'School'

    # â”€â”€ Export actions â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _get_headers(self):
        return [
            'First Name', 'Surname', 'Full Name',
            'WhatsApp Number', 'Country Code', 'Full WhatsApp',
            'Email', 'Age Range', 'Country', 'Region',
            'Category', 'Platform', 'Handle', 'Follower Range',
            'School Category', 'School Name', 'Level / Year',
            'Date Added', 'Days Since Added',
        ]

    def _get_row(self, contact):
        code = contact.country_code or ''
        num  = contact.whatsapp_number or ''
        full_wa = num if num.startswith('+') else (code + num)

        return [
            contact.first_name,
            contact.surname,
            contact.full_name,
            contact.whatsapp_number,
            contact.country_code or '',
            full_wa,
            contact.email or '',
            contact.age_range or '',
            contact.country or '',
            contact.region or '',
            str(contact.category) if contact.category else '',
            contact.platform or '',
            contact.handle or '',
            contact.follower_range or '',
            contact.get_school_category_display() if contact.school_category else '',
            contact.school_name or '',
            contact.level_year or '',
            contact.date_added.strftime('%Y-%m-%d %H:%M'),
            contact.days_since_added,
        ]

    def export_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SK Contact Base'

        header_fill = PatternFill(start_color='0C1F3F', end_color='0C1F3F', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        headers = self._get_headers()
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, contact in enumerate(queryset, start=2):
            for col_idx, value in enumerate(self._get_row(contact), start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col in ws.columns:
            col_cells = list(col)
            max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 40)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'SKContactBase_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        self.message_user(request, f'{queryset.count()} contacts exported to Excel.')
        return response

    export_to_excel.short_description = 'ðŸ“Š Export selected â†’ Excel'

    def export_to_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        filename = f'SKContactBase_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(self._get_headers())
        for contact in queryset:
            writer.writerow(self._get_row(contact))

        self.message_user(request, f'{queryset.count()} contacts exported to CSV.')
        return response

    export_to_csv.short_description = 'ðŸ“„ Export selected â†’ CSV'

from contacts.models import WhatsAppLog

@admin.register(WhatsAppLog)
class WhatsAppLogAdmin(admin.ModelAdmin):
    list_display  = ("phone", "template", "status", "timestamp", "contact")
    list_filter   = ("status", "template")
    search_fields = ("phone", "contact__first_name", "contact__surname")
    readonly_fields = ("contact", "template", "phone", "status", "error", "timestamp")
    ordering      = ("-timestamp",)



from contacts.models import Notification

@admin.register(Notification)
class NotificationAdmin(admin.ModelAdmin):
    list_display  = ("title", "type", "is_active", "created_at")
    list_filter   = ("type", "is_active")
    search_fields = ("title", "body")
    actions       = ["activate", "deactivate"]

    @admin.action(description="Activate selected notifications")
    def activate(self, request, queryset):
        queryset.update(is_active=True)

    @admin.action(description="Deactivate selected notifications")
    def deactivate(self, request, queryset):
        queryset.update(is_active=False)



from contacts.models import BulkMessage, ServiceRequest

@admin.register(BulkMessage)
class BulkMessageAdmin(admin.ModelAdmin):
    list_display  = ("template", "status", "sent_count", "failed_count", "created_by", "created_at")
    list_filter   = ("status", "template")
    readonly_fields = ("template", "filter_params", "sent_count", "failed_count", "created_by", "created_at")
    ordering      = ("-created_at",)

@admin.register(ServiceRequest)
class ServiceRequestAdmin(admin.ModelAdmin):
    list_display  = ("requester_name", "email", "phone", "status", "submitted_at")
    list_filter   = ("status",)
    search_fields = ("requester_name", "email", "phone")
    readonly_fields = ("requester_name", "email", "phone", "filter_criteria", "budget", "notes", "submitted_at")
    ordering      = ("-submitted_at",)



@admin.register(APIKey)
class APIKeyAdmin(admin.ModelAdmin):
    list_display = ('name', 'is_active', 'use_count', 'last_used', 'created_at')
    list_filter  = ('is_active',)
    readonly_fields = ('key_hash', 'use_count', 'last_used', 'created_at')

    def save_model(self, request, obj, form, change):
        if not change:
            obj_saved, raw = APIKey.create(obj.name)
            from django.contrib import messages as _msg
            _msg.warning(request, f'Copy this key now, it will not be shown again: {raw}')
            return
        super().save_model(request, obj, form, change)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# CATEGORY CHANGE REQUEST ADMIN
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
from contacts.models import CategoryChangeRequest
from django.utils import timezone as _tz

@admin.register(CategoryChangeRequest)
class CategoryChangeRequestAdmin(admin.ModelAdmin):

    list_display = [
        'whatsapp_link',
        'full_name',
        'current_category',
        'requested_category',
        'reason_short',
        'status_badge',
        'submitted_at',
    ]

    list_filter  = ['status', 'requested_category']

    search_fields = ['whatsapp_number', 'full_name', 'current_category']

    ordering = ['-submitted_at']

    readonly_fields = ['whatsapp_number', 'full_name', 'current_category',
                       'requested_category', 'reason', 'submitted_at', 'resolved_at']

    actions = ['mark_done', 'mark_rejected', 'apply_category_change']

    fieldsets = (
        ('Request Details', {
            'fields': ('whatsapp_number', 'full_name', 'current_category',
                       'requested_category', 'reason'),
        }),
        ('Status', {
            'fields': ('status', 'submitted_at', 'resolved_at'),
        }),
    )

    # â”€â”€ Custom columns â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def whatsapp_link(self, obj):
        clean = obj.whatsapp_number.replace('+', '').replace(' ', '')
        return format_html(
            '<a href="https://wa.me/{}" target="_blank" '
            'style="color:#25D366;font-weight:600;white-space:nowrap;">'
            'ðŸ’¬ {}</a>',
            clean, obj.whatsapp_number
        )
    whatsapp_link.short_description = 'WhatsApp'

    def reason_short(self, obj):
        if not obj.reason:
            return 'â€”'
        return obj.reason[:60] + ('â€¦' if len(obj.reason) > 60 else '')
    reason_short.short_description = 'Reason'

    def status_badge(self, obj):
        colours = {
            'pending':  ('#FEF3C7', '#B45309'),
            'done':     ('#DCFCE7', '#15803D'),
            'rejected': ('#FEE2E2', '#B91C1C'),
        }
        bg, fg = colours.get(obj.status, ('#F1F5F9', '#334155'))
        return format_html(
            '<span style="background:{};color:{};padding:2px 10px;border-radius:100px;'
            'font-size:11px;font-weight:700;">{}</span>',
            bg, fg, obj.get_status_display()
        )
    status_badge.short_description = 'Status'

    # â”€â”€ Bulk actions â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    @admin.action(description='âœ… Mark selected as Done')
    def mark_done(self, request, queryset):
        updated = queryset.update(status='done', resolved_at=_tz.now())
        self.message_user(request, f'{updated} request(s) marked as Done.')

    @admin.action(description='âŒ Mark selected as Rejected')
    def mark_rejected(self, request, queryset):
        updated = queryset.update(status='rejected', resolved_at=_tz.now())
        self.message_user(request, f'{updated} request(s) marked as Rejected.')

    @admin.action(description='ðŸ”„ Apply category change to Contact record')
    def apply_category_change(self, request, queryset):
        applied = 0
        skipped = 0
        for req in queryset.filter(status='pending'):
            try:
                contact = Contact.objects.get(whatsapp_number=req.whatsapp_number)
                contact.category = req.requested_category
                contact.save(update_fields=['category'])
                req.status      = 'done'
                req.resolved_at = _tz.now()
                req.save(update_fields=['status', 'resolved_at'])
                applied += 1
            except Contact.DoesNotExist:
                skipped += 1
        msg = f'{applied} category change(s) applied.'
        if skipped:
            msg += f' {skipped} skipped (contact not found by WhatsApp number).'
        self.message_user(request, msg)
