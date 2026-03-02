"""
SK Contact Base - Automation Script
This script handles:
1. SMS auto-messaging after contact save (via Africa's Talking)
2. Auto-sync to Google Drive
3. Daily statistics updates
"""

import os
import sys
import django
from datetime import datetime, timedelta
import json
from pathlib import Path

# Setup Django
sys.path.append(str(Path(__file__).parent.parent))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'sk_contact_base.settings')
django.setup()

from contacts.models import Contact, ContactStats, AutomationLog
from django.utils import timezone


# ─────────────────────────────────────────────
# AFRICA'S TALKING CREDENTIALS
# Set these in your .env or directly here for testing
# ─────────────────────────────────────────────
AT_USERNAME  = os.getenv('AT_USERNAME', 'sandbox')       # Use 'sandbox' for testing
AT_API_KEY   = os.getenv('AT_API_KEY',  'your_api_key_here')
AT_SENDER_ID = os.getenv('AT_SENDER_ID', '')             # Optional - leave blank in sandbox


class SMSAutomation:
    """
    Africa's Talking SMS Integration
    Install: pip install africastalking
    Docs:    https://developers.africastalking.com/docs/sms/sending
    """

    def __init__(self):
        self.username  = AT_USERNAME
        self.api_key   = AT_API_KEY
        self.sender_id = AT_SENDER_ID
        self._sms      = None

    def _get_gateway(self):
        """Lazily initialise the Africa's Talking SMS gateway."""
        if self._sms is None:
            import africastalking
            africastalking.initialize(self.username, self.api_key)
            self._sms = africastalking.SMS
        return self._sms

    def _format_phone(self, phone: str) -> str:
        """Ensure phone is in international format e.g. +233XXXXXXXXX"""
        phone = phone.strip().replace(' ', '').replace('-', '')
        if not phone.startswith('+'):
            phone = '+' + phone
        return phone

    # ── MESSAGE TEMPLATES ─────────────────────────────────────────────────
    # These are TEST messages. Swap with real content from your client later.
    # Only update _welcome_message() and _followup_message() when ready.

    def _welcome_message(self, contact) -> str:
        return (
            f"Hi {contact.first_name}! "
            f"Welcome to SK Network. "
            f"Your details have been received successfully. "
            f"We have added you as a {contact.get_category_display()} "
            f"on {contact.get_social_media_platform_display()}. "
            f"Our team will be in touch shortly. - SK Team"
        )

    def _followup_message(self, contact) -> str:
        return (
            f"Hello {contact.first_name}! "
            f"Stay connected with SK! "
            f"Join our community: "
            f"WhatsApp: https://wa.me/c/TEST_LINK | "
            f"Instagram: https://instagram.com/TEST_SKPAGE | "
            f"TikTok: https://tiktok.com/@TEST_SKPAGE. "
            f"More opportunities coming your way! - SK Team"
        )

    # ── SEND METHODS ───────────────────────────────────────────────────────

    def _send(self, phone: str, message: str) -> bool:
        """Core send method. Returns True on success."""
        sms = self._get_gateway()
        kwargs = {"message": message, "recipients": [phone]}
        if self.sender_id:
            kwargs["senderId"] = self.sender_id

        response = sms.send(**kwargs)
        recipients = response.get("SMSMessageData", {}).get("Recipients", [])
        return any(r.get("status") == "Success" for r in recipients)

    def send_welcome_message(self, contact) -> bool:
        """Send instant welcome SMS when a contact submits their details."""
        if not self.api_key or self.api_key == 'your_api_key_here':
            print("⚠️  Africa's Talking API key not set. Update AT_API_KEY.")
            return False

        phone = self._format_phone(contact.whatsapp_contact)

        try:
            success = self._send(phone, self._welcome_message(contact))

            if success:
                contact.whatsapp_message_sent = True
                contact.whatsapp_message_date = timezone.now()
                contact.save(update_fields=['whatsapp_message_sent', 'whatsapp_message_date'])

                AutomationLog.objects.create(
                    contact=contact,
                    action_type='sms',
                    status='success',
                    details="Welcome SMS sent via Africa's Talking"
                )
                print(f"✅ Welcome SMS sent to {contact.full_name} ({phone})")
                return True
            else:
                AutomationLog.objects.create(
                    contact=contact,
                    action_type='sms',
                    status='failed',
                    details="Africa's Talking returned non-success status"
                )
                print(f"❌ Welcome SMS failed for {contact.full_name}")
                return False

        except Exception as e:
            AutomationLog.objects.create(
                contact=contact,
                action_type='sms',
                status='error',
                details=str(e)
            )
            print(f"❌ Exception sending SMS: {str(e)}")
            return False

    def send_followup_message(self, contact) -> bool:
        """Send follow-up SMS with social links."""
        phone = self._format_phone(contact.whatsapp_contact)

        try:
            success = self._send(phone, self._followup_message(contact))

            if success:
                AutomationLog.objects.create(
                    contact=contact,
                    action_type='sms_followup',
                    status='success',
                    details="Follow-up SMS sent via Africa's Talking"
                )
                print(f"✅ Follow-up SMS sent to {contact.full_name} ({phone})")
                return True
            else:
                print(f"❌ Follow-up SMS failed for {contact.full_name}")
                return False

        except Exception as e:
            print(f"❌ Exception sending follow-up SMS: {str(e)}")
            return False

    def send_batch_messages(self, contact_ids) -> tuple:
        """Send welcome SMS to multiple contacts. Returns (success_count, total)."""
        contacts = Contact.objects.filter(id__in=contact_ids)
        success_count = sum(self.send_welcome_message(c) for c in contacts)
        return success_count, len(contact_ids)


class GoogleDriveAutomation:
    """Google Drive Integration for automatic weekly backup."""

    def __init__(self):
        self.credentials_file = os.getenv('GOOGLE_CREDENTIALS_FILE', 'credentials.json')
        self.folder_id = os.getenv('GOOGLE_DRIVE_FOLDER_ID', '')

    def upload_contacts_backup(self) -> bool:
        try:
            from googleapiclient.discovery import build
            from googleapiclient.http import MediaFileUpload
            from google.oauth2.service_account import Credentials
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            if not os.path.exists(self.credentials_file):
                print(f"⚠️  Google credentials not found at: {self.credentials_file}")
                return False

            credentials = Credentials.from_service_account_file(
                self.credentials_file,
                scopes=['https://www.googleapis.com/auth/drive.file']
            )
            service = build('drive', 'v3', credentials=credentials)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "SK Contacts"

            headers = [
                'First Name', 'Surname', 'Phone', 'Email', 'Category',
                'Age', 'Country', 'Platform', 'Followers', 'Date Added'
            ]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            for row, contact in enumerate(Contact.objects.all(), start=2):
                ws.cell(row=row, column=1, value=contact.first_name)
                ws.cell(row=row, column=2, value=contact.surname)
                ws.cell(row=row, column=3, value=contact.whatsapp_contact)
                ws.cell(row=row, column=4, value=contact.email or '')
                ws.cell(row=row, column=5, value=contact.get_category_display())
                ws.cell(row=row, column=6, value=contact.age_category)
                ws.cell(row=row, column=7, value=contact.country)
                ws.cell(row=row, column=8, value=contact.get_social_media_platform_display())
                ws.cell(row=row, column=9, value=contact.followers_count)
                ws.cell(row=row, column=10, value=contact.date_added.strftime('%Y-%m-%d'))

            filename = f"SK_Contacts_Backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"/tmp/{filename}"
            wb.save(filepath)

            file_metadata = {
                'name': filename,
                'parents': [self.folder_id] if self.folder_id else []
            }
            media = MediaFileUpload(
                filepath,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id, webViewLink'
            ).execute()

            Contact.objects.all().update(synced_to_drive=True)
            os.remove(filepath)

            print(f"✅ Backup uploaded: {file.get('webViewLink')}")
            return True

        except Exception as e:
            print(f"❌ Google Drive sync failed: {str(e)}")
            return False


class DailyStatsUpdater:
    @staticmethod
    def update_stats():
        today = timezone.now().date()
        total   = Contact.objects.count()
        today_n = Contact.objects.filter(date_added__date=today).count()

        stats, created = ContactStats.objects.update_or_create(
            date=today,
            defaults={'total_contacts': total, 'contacts_added_today': today_n}
        )
        action = "Created" if created else "Updated"
        print(f"✅ {action} stats: {total} total, {today_n} added today")
        return stats


def process_new_contacts():
    """Main cron entry point."""
    print("\n🤖 SK Contact Base - Automation Running...")
    print("=" * 50)

    cutoff_time  = timezone.now() - timedelta(hours=24)
    new_contacts = Contact.objects.filter(
        whatsapp_message_sent=False,
        date_added__gte=cutoff_time
    )

    if new_contacts.exists():
        print(f"\n📱 Sending SMS to {new_contacts.count()} new contact(s)...")
        sms = SMSAutomation()
        for contact in new_contacts:
            sms.send_welcome_message(contact)
    else:
        print("\n✓ No new contacts to process")

    print("\n📊 Updating daily statistics...")
    DailyStatsUpdater.update_stats()

    if datetime.now().weekday() == 0:  # Every Monday
        print("\n☁️  Running weekly Google Drive backup...")
        GoogleDriveAutomation().upload_contacts_backup()

    print("\n✅ Automation complete!")
    print("=" * 50)


if __name__ == "__main__":
    process_new_contacts()


"""
════════════════════════════════════════════
SETUP FOR TESTING
════════════════════════════════════════════

Step 1 - Sign up at https://africastalking.com
Step 2 - Go to Sandbox → API Key → copy your key
Step 3 - Set in your terminal before running:

   Windows (PowerShell):
   $env:AT_USERNAME = "sandbox"
   $env:AT_API_KEY  = "your_sandbox_api_key"

   Mac/Linux:
   export AT_USERNAME=sandbox
   export AT_API_KEY=your_sandbox_api_key

Step 4 - Run a test:
   python automation.py

════════════════════════════════════════════
DJANGO SIGNAL — fires SMS the INSTANT a contact is saved
════════════════════════════════════════════

Create contacts/signals.py and add:

    from django.db.models.signals import post_save
    from django.dispatch import receiver
    from .models import Contact
    from automation import SMSAutomation

    @receiver(post_save, sender=Contact)
    def send_sms_on_new_contact(sender, instance, created, **kwargs):
        if created and not instance.whatsapp_message_sent:
            sms = SMSAutomation()
            sms.send_welcome_message(instance)

Then in contacts/apps.py inside ContactsConfig class add:

    def ready(self):
        import contacts.signals

════════════════════════════════════════════
SWAP REAL MESSAGES LATER
════════════════════════════════════════════
When your client provides the real content and social links,
update ONLY these two methods in SMSAutomation:
  → _welcome_message()
  → _followup_message()
Everything else stays the same.
════════════════════════════════════════════
"""